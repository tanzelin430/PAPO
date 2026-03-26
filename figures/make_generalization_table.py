"""Generate Generalization table Excel for pubtab conversion."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Generalization"

# ── Data ──
# (ORM, PA-GRPO) per benchmark; None = TBD
data = {
    "Qwen2.5-3B": {
        "OlympiadBench": (35.3, 38.6),
        "MATH-500":      (69.0, 73.2),
        "AIME'24":       (8.3, 10.0),
        "AIME'25":       (4.2, 3.3),
        "AIME'26":       (8.3, 9.5),
        "GPQA":          (38.8, 36.3),
        "HumanEval":     (39.3, 52.5),
    },
    "Qwen2.5-7B": {
        "OlympiadBench": (46.3, 51.3),
        "MATH-500":      (80.2, 82.3),
        "AIME'24":       (10.8, 15.8),
        "AIME'25":       (10.8, 13.1),
        "AIME'26":       (11.7, 12.5),
        "GPQA":          (40.7, 42.4),
        "HumanEval":     (61.6, 63.9),
    },
    "Qwen2.5-14B": {
        "OlympiadBench": (None, None),
        "MATH-500":      (None, None),
        "AIME'24":       (None, None),
        "AIME'25":       (None, None),
        "AIME'26":       (None, None),
        "GPQA":          (None, None),
        "HumanEval":     (None, None),
    },
    "Qwen3-4B": {
        "OlympiadBench": (None, None),
        "MATH-500":      (None, None),
        "AIME'24":       (None, None),
        "AIME'25":       (None, None),
        "AIME'26":       (None, None),
        "GPQA":          (None, None),
        "HumanEval":     (None, None),
    },
}

# Display order: 4 competition + 1 standard + 1 STEM + 1 code
bench_order = ["OlympiadBench", "AIME'24", "AIME'25", "AIME'26",
               "MATH-500", "GPQA", "HumanEval"]

# ── Styles ──
header_font = Font(bold=True, size=10)
data_font = Font(size=10)
bold_font = Font(bold=True, size=10)
green_font = Font(color="228B22", size=9, bold=True)  # improvement
red_font = Font(color="CC0000", size=9, bold=True)    # decline
gray_font = Font(color="999999", size=10, italic=True)
center = Alignment(horizontal="center", vertical="center")
header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(
    bottom=Side(style="thin", color="000000"),
)

# Column layout:
# A=Model, B=Method, C-F=Competition Math (4 cols), G=Standard, H=STEM, I=Code, J=Avg

# ── Header Row 1: Category grouping ──
ws.merge_cells("A1:B1")
ws.cell(row=1, column=1, value="").font = header_font

# Competition Math spans columns 3-6 (Olympiad, AIME'24, AIME'25, AIME'26)
ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)
ws.cell(row=1, column=3, value="Competition Mathematics").font = header_font
ws.cell(row=1, column=3).alignment = center
ws.cell(row=1, column=3).fill = header_fill

# Standard Math = col 7
ws.cell(row=1, column=7, value="Standard").font = header_font
ws.cell(row=1, column=7).alignment = center
ws.cell(row=1, column=7).fill = header_fill

# STEM = col 8
ws.cell(row=1, column=8, value="STEM").font = header_font
ws.cell(row=1, column=8).alignment = center
ws.cell(row=1, column=8).fill = header_fill

# Code = col 9
ws.cell(row=1, column=9, value="Code").font = header_font
ws.cell(row=1, column=9).alignment = center
ws.cell(row=1, column=9).fill = header_fill

# Avg = col 10
ws.cell(row=1, column=10, value="").font = header_font

# ── Header Row 2: Benchmark names ──
headers = ["Model", "Method",
           "Olympiad", "AIME'24", "AIME'25", "AIME'26",
           "MATH-500", "GPQA", "HumanEval", "Avg."]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border

# ── Data Rows ──
row = 3
for model_name, benchmarks_data in data.items():
    # ORM row
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
    ws.cell(row=row, column=1, value=model_name).font = bold_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=row, column=2, value="ORM").font = data_font
    ws.cell(row=row, column=2).alignment = center

    orm_vals = []
    pa_vals = []
    for col_idx, bench in enumerate(bench_order, 3):
        orm_v, pa_v = benchmarks_data[bench]
        orm_vals.append(orm_v)
        pa_vals.append(pa_v)

        if orm_v is not None:
            ws.cell(row=row, column=col_idx, value=orm_v).font = data_font
        else:
            ws.cell(row=row, column=col_idx, value="—").font = gray_font
        ws.cell(row=row, column=col_idx).alignment = center

    # ORM Avg (col = 3 + len(bench_order))
    avg_col = 3 + len(bench_order)
    valid_orm = [v for v in orm_vals if v is not None]
    if valid_orm:
        avg_orm = sum(valid_orm) / len(valid_orm)
        ws.cell(row=row, column=avg_col, value=round(avg_orm, 1)).font = data_font
    else:
        ws.cell(row=row, column=avg_col, value="—").font = gray_font
    ws.cell(row=row, column=avg_col).alignment = center

    # PA-GRPO row
    row += 1
    ws.cell(row=row, column=2, value="PA-GRPO").font = bold_font
    ws.cell(row=row, column=2).alignment = center

    for col_idx, (bench, orm_v, pa_v) in enumerate(zip(bench_order, orm_vals, pa_vals), 3):
        if pa_v is not None:
            cell = ws.cell(row=row, column=col_idx)
            if orm_v is not None and pa_v > orm_v:
                cell.value = pa_v
                cell.font = bold_font
            elif orm_v is not None and pa_v < orm_v:
                cell.value = pa_v
                cell.font = data_font
            else:
                cell.value = pa_v
                cell.font = bold_font
        else:
            ws.cell(row=row, column=col_idx, value="—").font = gray_font
        ws.cell(row=row, column=col_idx).alignment = center

    # PA-GRPO Avg
    valid_pa = [v for v in pa_vals if v is not None]
    if valid_pa:
        avg_pa = sum(valid_pa) / len(valid_pa)
        ws.cell(row=row, column=avg_col, value=round(avg_pa, 1)).font = bold_font
    else:
        ws.cell(row=row, column=avg_col, value="—").font = gray_font
    ws.cell(row=row, column=avg_col).alignment = center

    # Delta row
    row += 1
    ws.cell(row=row, column=2, value="Δ").font = Font(size=9, italic=True)
    ws.cell(row=row, column=2).alignment = center

    for col_idx, (orm_v, pa_v) in enumerate(zip(orm_vals, pa_vals), 3):
        if orm_v is not None and pa_v is not None:
            delta = pa_v - orm_v
            cell = ws.cell(row=row, column=col_idx)
            if delta > 0:
                cell.value = f"+{delta:.1f}"
                cell.font = green_font
            elif delta < 0:
                cell.value = f"{delta:.1f}"
                cell.font = red_font
            else:
                cell.value = "0.0"
                cell.font = data_font
        else:
            ws.cell(row=row, column=col_idx, value="—").font = gray_font
        ws.cell(row=row, column=col_idx).alignment = center

    # Delta Avg
    if valid_orm and valid_pa:
        delta_avg = sum(valid_pa) / len(valid_pa) - sum(valid_orm) / len(valid_orm)
        cell = ws.cell(row=row, column=avg_col)
        if delta_avg > 0:
            cell.value = f"+{delta_avg:.1f}"
            cell.font = green_font
        else:
            cell.value = f"{delta_avg:.1f}"
            cell.font = red_font
    else:
        ws.cell(row=row, column=avg_col, value="—").font = gray_font
    ws.cell(row=row, column=avg_col).alignment = center

    # Add border after each model group
    for c in range(1, avg_col + 1):
        ws.cell(row=row, column=c).border = thin_border

    row += 1

# ── Column widths ──
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
for col in range(3, avg_col + 1):
    ws.column_dimensions[get_column_letter(col)].width = 10

output_path = "/home/tanzelin-p/PSRO4math/figures/generalization_table.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
